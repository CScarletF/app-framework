"""
export.py -- generic multi-sheet Excel export, framework-level (not
module-specific). Every module currently registered with app.py (i.e.
every table in the shared `tables` dict) gets its own sheet, named after
its table. Adding a new module later needs no changes here -- this
iterates whatever's in `tables` at request time.
"""

from io import BytesIO
from flask import Blueprint, send_file
from sqlalchemy import select
from sqlalchemy.engine import Engine
from sqlalchemy import Table
from openpyxl import Workbook


def build_export_blueprint(tables: dict[str, Table], engine: Engine) -> Blueprint:
    bp = Blueprint("export_routes", __name__)

    @bp.get("/api/export/xlsx")
    def export_all_xlsx():
        wb = Workbook()
        # Workbook() ships with one default empty sheet -- drop it once we
        # know we're about to add real ones, so there's no stray blank tab.
        wb.remove(wb.active)

        with engine.connect() as conn:
            for table_name, table in tables.items():
                ws = wb.create_sheet(title=table_name[:31])  # Excel's sheet-name length limit
                columns = list(table.columns.keys())
                ws.append(columns)

                rows = conn.execute(select(table).order_by(table.c.id)).mappings().all()
                for row in rows:
                    ws.append([str(row[c]) if row[c] is not None else "" for c in columns])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="app_export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return bp